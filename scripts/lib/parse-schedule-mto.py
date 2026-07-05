#!/usr/bin/env python3
"""Parse SCHEDULE_MTO.xlsx truck delivery rows → JSON stdout."""
from __future__ import annotations

import json
import sys
from datetime import date, datetime

import openpyxl


def main() -> None:
    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        if not isinstance(row[0], (datetime, date)):
            continue
        truck_date = row[0].date().isoformat() if hasattr(row[0], "date") else str(row[0])[:10]
        rosd = row[4] if len(row) > 4 else None
        if isinstance(rosd, (datetime, date)):
            rosd = rosd.date().isoformat()
        rows.append(
            {
                "truck_date": truck_date,
                "truck_no": row[1],
                "units": row[2],
                "rosd": rosd,
            }
        )
    print(json.dumps(rows))


if __name__ == "__main__":
    main()
