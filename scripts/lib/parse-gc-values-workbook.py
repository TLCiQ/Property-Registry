#!/usr/bin/env python3
"""Parse GC Values Workbook / Install Sub Contract Workbook → delivery milestones JSON."""
from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

DATE_LABELS = {
    "plans dated": "Plans Dated",
    "plan dated": "Plans Dated",
    "shop drawing": "Shop Drawings Due",
    "submittal": "Submittals Due",
    "fabrication": "Fabrication Complete",
    "first delivery": "First Delivery (Truck 1)",
    "delivery start": "First Delivery (Truck 1)",
    "start delivery": "First Delivery (Truck 1)",
    "rosd": "ROSD",
    "required on site": "ROSD",
    "required on-site": "ROSD",
    "last delivery": "Last Delivery",
    "final delivery": "Last Delivery",
    "install start": "Install Start",
    "installation start": "Install Start",
    "install complete": "Install Complete",
    "installation complete": "Install Complete",
    "substantial completion": "Substantial Completion",
    "work commence": "Work Commencement (Contract)",
    "commencement": "Work Commencement (Contract)",
    "mobilization": "Mobilization",
}


def to_iso(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    if not s:
        return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
    if m:
        y = int(m.group(3))
        if y < 100:
            y += 2000 if y < 70 else 1900
        return date(y, int(m.group(1)), int(m.group(2))).isoformat()
    m = re.match(
        r"^(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{1,2}),?\s+(\d{4})$",
        s,
        re.I,
    )
    if m:
        return datetime.strptime(f"{m.group(1)} {m.group(2)}, {m.group(3)}", "%B %d, %Y").date().isoformat()
    return None


def label_for(text: str) -> str | None:
    t = text.lower().strip()
    for key, name in DATE_LABELS.items():
        if key in t:
            return name
    return None


def parse_sheet(ws) -> list[dict]:
    milestones: list[dict] = []
    seen: set[str] = set()

    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c is not None and str(c).strip()]
        if len(cells) < 2:
            continue

        # Label in col0, date in col1 (or scan row for date)
        label_text = str(cells[0]).strip()
        label = label_for(label_text)
        iso = None
        for c in cells[1:8]:
            iso = to_iso(c)
            if iso:
                break
        if label and iso:
            key = f"{label}|{iso}"
            if key not in seen:
                seen.add(key)
                milestones.append(
                    {
                        "milestone_name": label,
                        "target_date": iso,
                        "milestone_category": "delivery"
                        if "delivery" in label.lower() or label == "ROSD"
                        else "install"
                        if "install" in label.lower() or "commencement" in label.lower()
                        else "contract",
                        "source_field_path": f"gc_values_workbook:{label_text[:40]}",
                        "status": "pending",
                    }
                )

        # Truck schedule rows: date | truck# | units
        if isinstance(cells[0], (datetime, date)) or re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", str(cells[0])):
            truck_date = to_iso(cells[0])
            if not truck_date:
                continue
            truck_no = cells[1] if len(cells) > 1 else None
            if truck_no == "Truck" or str(truck_no).lower() == "truck":
                continue
            if "First Delivery (Truck 1)" not in seen and isinstance(truck_no, (int, float)):
                milestones.append(
                    {
                        "milestone_name": "First Delivery (Truck 1)",
                        "target_date": truck_date,
                        "milestone_category": "delivery",
                        "source_field_path": "gc_values_workbook:truck_1",
                        "status": "pending",
                    }
                )
                seen.add("First Delivery (Truck 1)")

    # Last truck delivery if we saw truck rows
    truck_dates = [
        m["target_date"]
        for m in milestones
        if m["source_field_path"].startswith("gc_values_workbook:truck")
        or "Delivery" in m["milestone_name"]
    ]
    if truck_dates:
        last = max(truck_dates)
        if "Last Delivery" not in seen:
            milestones.append(
                {
                    "milestone_name": "Last Delivery",
                    "target_date": last,
                    "milestone_category": "delivery",
                    "source_field_path": "gc_values_workbook:last_truck",
                    "status": "pending",
                }
            )

    return milestones


def main() -> None:
    path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_ms: list[dict] = []
    seen: set[str] = set()
    for sn in wb.sheetnames:
        for m in parse_sheet(wb[sn]):
            key = m["milestone_name"].lower()
            if key not in seen:
                seen.add(key)
                all_ms.append(m)
    all_ms.sort(key=lambda x: x["target_date"])
    print(json.dumps({"milestones": all_ms, "sheets": wb.sheetnames}))


if __name__ == "__main__":
    main()
