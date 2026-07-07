#!/usr/bin/env python3
"""Generic BSI unit matrix extract from Box xlsx → JSON for Registry-iQ structure ingest."""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl


def norm_unit(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in {"unit #", "unit#", "unit"}:
        return None
    if s.lower().startswith("level "):
        return None
    try:
        if isinstance(val, (int, float)) and float(val).is_integer():
            return str(int(float(val)))
    except (TypeError, ValueError):
        pass
    if s.upper().startswith("TIER "):
        return re.sub(r"\s+", " ", s.upper().replace("TIER ", "TIER ", 1)).replace("TIER  ", "TIER ")
    return s


def infer_beds(unit_type: str | None) -> int:
    if not unit_type:
        return 0
    t = unit_type.strip().upper()
    m = re.match(r"^B(\d+)", t)
    if m:
        return int(m.group(1))
    m = re.match(r"^D(\d+(?:\.\d+)?)", t)
    if m:
        return int(float(m.group(1)))
    if t.startswith("S"):
        return 0
    if t.startswith("A") or t.startswith("E"):
        return 1
    return 0


def score_sheet(headers: list[str]) -> int:
    lower = [h.lower() for h in headers]
    score = 0
    if any("unit" in h and "#" in h for h in lower):
        score += 5
    if any("unit type" in h or h == "type" for h in lower):
        score += 3
    if any("level" in h for h in lower):
        score += 1
    if any("all building" in h for h in headers):
        score += 1
    return score


def col(col_idx, *names):
    for n in names:
        for k, i in col_idx.items():
            if n.lower() in k.lower():
                return i
    return None


def detect_sheet(ws) -> tuple[int, dict[str, int], str | None] | None:
    best = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        headers = [str(c).strip() if c is not None else "" for c in row]
        sc = score_sheet(headers)
        if sc >= 5 and (best is None or sc > best[0]):
            col_idx = {headers[i]: i for i in range(len(headers)) if headers[i]}
            best = (sc, row_idx, col_idx, headers)
    if not best:
        return None
    _, header_row, col_idx, headers = best
    unit_col = col(col_idx, "Unit #", "Unit#", "Unit")
    if unit_col is None:
        return None
    return header_row, col_idx, None


def extract_sheet(ws, sheet_name: str) -> list[dict]:
    info = detect_sheet(ws)
    if not info:
        return []
    header_row, col_idx, _ = info
    unit_col = col(col_idx, "Unit #", "Unit#", "Unit")
    type_col = col(col_idx, "Unit Type", "Type", "UnitType")
    level_col = col(col_idx, "Level", "Floor")
    area_col = col(col_idx, "Area", "Construction Area", "Sq Ft", "Sqft")

    units = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if unit_col is None or unit_col >= len(row):
            continue
        unit = norm_unit(row[unit_col])
        if not unit:
            continue
        unit_type = (
            str(row[type_col]).strip()
            if type_col is not None and type_col < len(row) and row[type_col]
            else None
        )
        level = (
            str(row[level_col]).strip()
            if level_col is not None and level_col < len(row) and row[level_col] is not None
            else None
        )
        area = (
            str(row[area_col]).strip()
            if area_col is not None and area_col < len(row) and row[area_col] is not None
            else None
        )
        units.append(
            {
                "unit_number": unit,
                "unit_type": unit_type,
                "level": level,
                "construction_area": area,
                "sheet": sheet_name,
            }
        )
    return units


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--project-id", default="")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        raise SystemExit(f"Matrix not found: {xlsx}")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    all_units: dict[str, dict] = {}
    type_counts: Counter[str] = Counter()
    sheets_used: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        extracted = extract_sheet(ws, sheet_name)
        if not extracted:
            continue
        sheets_used.append(sheet_name)
        for u in extracted:
            all_units[u["unit_number"]] = u
            if u.get("unit_type"):
                type_counts[u["unit_type"]] += 1

    wb.close()

    unit_types = []
    for name, count in sorted(type_counts.items()):
        unit_types.append(
            {
                "unit_type_name": name,
                "unit_count": count,
                "beds_per_unit": infer_beds(name),
                "standard_bedrooms": infer_beds(name),
                "bathrooms": 0,
            }
        )

    units = list(all_units.values())
    out = {
        "source": str(xlsx),
        "project_id": args.project_id,
        "sheets": sheets_used,
        "unit_count": len(units),
        "unit_type_count": len(unit_types),
        "units": units,
        "unit_types": unit_types,
    }
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, indent=2))
    print(
        f"Wrote {out_path} — {out['unit_count']} units, {out['unit_type_count']} types "
        f"from {len(sheets_used)} sheet(s)"
    )


if __name__ == "__main__":
    main()
