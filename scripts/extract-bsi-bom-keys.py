#!/usr/bin/env python3
"""Extract per-unit kitchen BOM keys (MW + THUS/OPP) from BSI matrix xlsx."""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl


def norm_tier(val) -> str | None:
    if val is None or val == "":
        return None
    s = str(val).strip()
    if not s or s == ".":
        return None
    if s.upper().startswith("TIER"):
        s = re.sub(r"\s+", " ", s.upper())
        s = re.sub(r"\s+-", "-", s)
        s = re.sub(r"-\s+", "-", s)
        return s
    return None


def norm_unit(val, *, allow_tier: bool = True) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)) and float(val).is_integer():
        return str(int(float(val)))
    s = str(val).strip()
    if not s or s in {".", "-"} or s.lower().startswith("level "):
        return None
    if s.upper().startswith("TIER"):
        return norm_tier(s) if allow_tier else None
    return s


def parse_top_opp(raw) -> str | None:
    if raw is None or raw == "":
        return None
    s = str(raw).strip().upper()
    if not s or s == ".":
        return None
    if s.startswith("T") or s == "THUS":
        return "THUS"
    if s.startswith("O") or s == "OPP":
        return "OPP"
    return None


def norm_mw(raw) -> str | None:
    if raw is None or raw == "":
        return None
    s = str(raw).strip().upper()
    if not s or s in {".", "TRUE", "FALSE"}:
        return None
    if s.startswith("MUR_"):
        s = s[4:]
    m = re.match(r"^(MW[\d.]+)", s)
    if not m:
        return None
    code = m.group(1)
    m2 = re.match(r"^MW(\d+(?:\.\d+)?)([A-Z]?)$", code, re.I)
    if not m2:
        return code
    num, suffix = m2.group(1), m2.group(2)
    if "." in num:
        a, b = num.split(".", 1)
        num = f"{int(a):02d}.{b}"
    else:
        num = f"{int(num):02d}"
    return f"MW{num}{suffix}"


def resolve_mw_col(headers: list[str]) -> str | None:
    lower = [h.lower().strip() for h in headers]
    for i, h in enumerate(lower):
        if h == "mw":
            return headers[i]
    for prefer in (
        "shop drawing",
        "kitchen sd",
        "drawing",
        "kitchen cab",
        "kitchen",
    ):
        for i, h in enumerate(lower):
            if prefer in h and "counter" not in h and "top" not in h and "type" not in h:
                return headers[i]
    return None


def resolve_kitchen_type_col(headers: list[str]) -> str | None:
    lower = [h.lower() for h in headers]
    for i, h in enumerate(lower):
        if h in {"kitchen type", "kitchen_type"}:
            return headers[i]
    return None


def score_sheet(headers: list[str]) -> int:
    lower = [h.lower() for h in headers]
    score = 0
    if any("unit #" in h or h.strip() == "unit #" for h in lower):
        score += 5
    if any(h.strip() == "tier" for h in lower):
        score += 5
    if resolve_mw_col(headers) or resolve_kitchen_type_col(headers):
        score += 4
    if any("thus" in h or "t/o" in h or "t/opp" in h for h in lower):
        score += 3
    if any("unit type" in h for h in lower):
        score += 2
    if any("all building" in h for h in headers):
        score += 1
    return score


def header_index(headers: list[str], *needles: str) -> int | None:
    lower = [h.lower() for h in headers]
    for needle in needles:
        for i, h in enumerate(lower):
            if needle in h:
                return i
    return None


def detect_layout(ws, sheet_name: str) -> dict | None:
    """Return column layout for a matrix sheet, or None if not a BOM sheet."""
    best = None
    rows_cache: list[tuple] = []
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=14, values_only=True), 1):
        rows_cache.append(row)
        headers = [str(c).strip() if c is not None else "" for c in row]
        sc = score_sheet(headers)
        if "review" in sheet_name.lower():
            sc -= 3
        if sc >= 9 and (best is None or sc > best[0]):
            best = (sc, ri, headers)

    if not best:
        return None

    _, header_row, headers = best
    col = {headers[i]: i for i in range(len(headers)) if headers[i]}

    tier_col = header_index(headers, "tier")
    unit_type_col = header_index(headers, "unit type")
    unit_col = header_index(headers, "unit #", "unit#")
    thus_col = header_index(headers, "thus", "t/o", "t/opp")
    mw_header = resolve_mw_col(headers)
    kitchen_type_header = resolve_kitchen_type_col(headers)
    mw_col = col.get(mw_header) if mw_header else None
    kitchen_type_col = col.get(kitchen_type_header) if kitchen_type_header else None
    type_col = unit_type_col

    # Clemson-style: TIER column + UNIT TYPE — join registry by unit type name
    if tier_col is not None and unit_type_col is not None and unit_col is None:
        unit_col = unit_type_col
        mode = "unit_type_key"
        if thus_col is None:
            thus_col = header_index(headers, "t/o")
        if mw_col is None:
            for i, h in enumerate(headers):
                if h.strip().upper() == "MW":
                    mw_col = i
                    break
    else:
        mode = "standard"

    # Tampa-style two-row header: KITCHEN spans T/OPP + MW on next row
    subheader_row = header_row + 1 if header_row < len(rows_cache) else None
    if subheader_row and unit_col is not None:
        sub = rows_cache[subheader_row - 1]
        sub_vals = [str(c).strip().upper() if c is not None else "" for c in sub]
        kitchen_start = header_index(headers, "kitchen")
        if kitchen_start is not None and thus_col is None:
            if kitchen_start < len(sub_vals) and sub_vals[kitchen_start] in {"T/OPP", "T/O", "THUS/OPP"}:
                thus_col = kitchen_start
                if kitchen_start + 1 < len(sub_vals) and sub_vals[kitchen_start + 1] == "MW":
                    mw_col = kitchen_start + 1
                mode = "tier_as_unit"

    # ALL BUILDINGS offset: BUILDIND | Unit # | ...
    if unit_col is None and header_index(headers, "unit #", "unit#") is not None:
        unit_col = header_index(headers, "unit #", "unit#")

    if unit_col is None:
        return None

    # Tier values in UNIT # column (25337)
    sample_vals = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 8, values_only=True):
        if unit_col < len(row) and row[unit_col]:
            sample_vals.append(str(row[unit_col]).strip().upper())
    if sample_vals and all(v.startswith("TIER") for v in sample_vals if v and v != "."):
        mode = "tier_as_unit"

    if thus_col is None or (mw_col is None and kitchen_type_col is None):
        return None

    return {
        "sheet": sheet_name,
        "header_row": header_row,
        "mode": mode,
        "unit_col": unit_col,
        "type_col": type_col,
        "thus_col": thus_col,
        "mw_col": mw_col,
        "kitchen_type_col": kitchen_type_col,
        "tier_col": tier_col,
    }


def extract_sheet(ws, sheet_name: str) -> list[dict]:
    layout = detect_layout(ws, sheet_name)
    if not layout:
        return []

    units = []
    for row in ws.iter_rows(min_row=layout["header_row"] + 1, values_only=True):
        unit_col = layout["unit_col"]
        if unit_col >= len(row):
            continue

        allow_tier = layout["mode"] in {"tier_as_unit", "unit_type_key"}
        unit_number = norm_unit(row[unit_col], allow_tier=allow_tier or layout["mode"] == "tier_as_unit")
        if not unit_number:
            continue

        thus_col = layout["thus_col"]
        top_opp = parse_top_opp(row[thus_col] if thus_col is not None and thus_col < len(row) else None)

        mw_col = layout["mw_col"]
        kitchen_type_col = layout["kitchen_type_col"]
        kitchen_raw = None
        mw_base = None

        if mw_col is not None and mw_col < len(row):
            kitchen_raw = row[mw_col]
            mw_base = norm_mw(kitchen_raw)

        if not mw_base and kitchen_type_col is not None and kitchen_type_col < len(row):
            kt = row[kitchen_type_col]
            if kt not in (None, ""):
                kitchen_raw = kt
                mw_base = None  # resolved at ingest via kit map / unit type

        if not top_opp or (not mw_base and kitchen_raw in (None, "", ".")):
            continue

        type_col = layout["type_col"]
        tier_col = layout["tier_col"]
        unit_type_name = (
            str(row[type_col]).strip()
            if type_col is not None and type_col < len(row) and row[type_col]
            else None
        )
        tier_name = (
            norm_tier(row[tier_col])
            if tier_col is not None and tier_col < len(row)
            else None
        )

        entry = {
            "unit_number": unit_number,
            "unit_type_name": unit_type_name,
            "thus_opp": top_opp,
            "kitchen_cab": str(kitchen_raw).strip() if kitchen_raw else None,
            "mw_base": mw_base,
            "sheet": sheet_name,
            "join_mode": layout["mode"],
        }
        if tier_name:
            entry["tier_name"] = tier_name
        if mw_base is None and kitchen_raw is not None:
            entry["kitchen_type"] = str(kitchen_raw).strip()
        units.append(entry)

    return units


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        print(json.dumps({"error": f"Matrix not found: {xlsx}"}))
        raise SystemExit(1)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    all_units: dict[str, dict] = {}
    sheets_used = []

    for sn in wb.sheetnames:
        if "review" in sn.lower():
            continue
        ws = wb[sn]
        extracted = extract_sheet(ws, sn)
        if extracted:
            sheets_used.append(sn)
            for u in extracted:
                all_units[u["unit_number"]] = u

    wb.close()
    units = list(all_units.values())
    payload = {
        "source": str(xlsx),
        "sheets": sheets_used,
        "units": units,
        "count": len(units),
    }
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=2))
    print(f"Wrote {out} — {len(units)} units w/ BOM keys from {len(sheets_used)} sheet(s)")


if __name__ == "__main__":
    main()
