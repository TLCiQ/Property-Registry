#!/usr/bin/env python3
"""Extract vanity (and vanity-adjacent) SKUs from Morgan Hill Counts Workbook MW tabs."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_DIR = Path(
    "/Users/geoffreyjackson/Library/CloudStorage/Box-Box/Team Folder/Projects/"
    "25048 - Carrollton, TX - Morgan Hill/PROJECT MANAGING/SHOP DRAWINGS/Cabinets/Counts Workbook"
)
OUT = ROOT / ".firecrawl" / "mh-vanity-bom.json"

VANITY_SKU = re.compile(r"^(FHVSB|VF\d|VB09|VDB)", re.I)
KNOWN_VANITY = {
    "FHVSB31REM", "FHVSB33REM", "FHVSB36REM", "FHVSB30", "FHVSB42REM",
    "VF1", "VF3", "VB09L", "VB09R", "VDB12-3",
}


def parse_tab_name(name: str) -> tuple[str, str, str] | None:
    m = re.match(r"^(MW[\d.]+)\s*-\s*(THUS|OPP)\s*$", name.strip(), re.I)
    if not m:
        return None
    return m.group(1), m.group(2).upper(), name


def scheme_from_file(path: Path) -> str:
    return "2" if "Scheme 2" in path.name else "1"


def is_vanity_sku(sku: str) -> bool:
    sku = sku.strip().upper()
    return sku in {s.upper() for s in KNOWN_VANITY} or bool(VANITY_SKU.match(sku))


def extract_workbook(path: Path) -> list[dict]:
    scheme = scheme_from_file(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lines: list[dict] = []
    for sheet in wb.sheetnames:
        parsed = parse_tab_name(sheet)
        if not parsed:
            continue
        mw_base, top_opp, tab = parsed
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=1, max_row=400, values_only=True):
            if not row or len(row) < 5:
                continue
            qty = row[2]
            sku = row[4]
            if sku is None or not isinstance(sku, str):
                continue
            sku = sku.strip()
            if not sku or not is_vanity_sku(sku):
                continue
            try:
                qpu = float(qty) if qty is not None else 1.0
            except (TypeError, ValueError):
                qpu = 1.0
            lines.append(
                {
                    "scheme": scheme,
                    "mw_base": mw_base,
                    "top_opp": top_opp,
                    "sku": sku,
                    "qty_per_unit": qpu,
                    "sku_role": "vanity",
                    "source_tab": tab,
                    "source_file": path.name,
                }
            )
    wb.close()
    return lines


def main() -> None:
    workbooks = sorted(WORKBOOK_DIR.glob("Carrolton, TX - Morgan Hill Workbook Scheme *.xlsm"))
    if not workbooks:
        print(json.dumps({"error": f"No workbooks in {WORKBOOK_DIR}"}))
        sys.exit(1)
    all_lines: list[dict] = []
    for wb_path in workbooks:
        all_lines.extend(extract_workbook(wb_path))
    index: dict[str, list] = {}
    for line in all_lines:
        key = f"{line['scheme']}|{line['mw_base']}|{line['top_opp']}"
        index.setdefault(key, []).append(line)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({"lines": all_lines, "keys": len(index), "count": len(all_lines)}, indent=2))
    print(f"Wrote {OUT} — {len(all_lines)} vanity lines, {len(index)} BOM keys")


if __name__ == "__main__":
    main()
