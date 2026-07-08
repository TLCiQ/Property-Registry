#!/usr/bin/env python3
"""Extract CSL Sales workbook lines for PH (phantom) Registry ingest.

Supports:
  - quote_rev: sheets named Quote Rev.* or Quote - * (Sales Project Workbooks / CSL)
  - asia_po: SST purchasing Form sheet (SKU NUMBER / QTY ORDER)
  - sage_invoice_pdf: Sage order confirmation PDF (interim when Sales xlsx missing)

Stdout: JSON { format, header, lines[], unit_mix[], warnings[] }
"""
from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl required"}))
    sys.exit(1)


def txt(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def num(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        n = float(v)
        return n if n == n else None
    except (TypeError, ValueError):
        return None


def find_quote_sheets(names: list[str]) -> list[str]:
    out = []
    for n in names:
        low = n.lower().strip()
        if low.startswith("quote rev") or low.startswith("quote -"):
            out.append(n)
    return out


def parse_quote_sheet(ws, sheet_name: str) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = -1
    for i, row in enumerate(rows[:40]):
        vals = list(row or [])
        if len(vals) > 6 and txt(vals[2]) == "SKU" and txt(vals[6]) == "QTY":
            header_idx = i
            break
    if header_idx < 0:
        raise ValueError(f"{sheet_name}: SKU/QTY header not found")

    header = {"project_name": None, "developer": None, "quote_sheet": sheet_name}
    for i in range(header_idx):
        r = rows[i] or []
        a = txt(r[0] if len(r) > 0 else None)
        if a and a.startswith("Project Name"):
            header["project_name"] = txt(r[2] if len(r) > 2 else None)
        if a and a.startswith("Developer"):
            header["developer"] = txt(r[2] if len(r) > 2 else None)

    lines = []
    unit_mix = []
    for i in range(header_idx + 1, len(rows)):
        r = list(rows[i] or [])
        if len(r) > 7 and txt(r[7]) == "Subtotal":
            break
        sku = txt(r[2] if len(r) > 2 else None)
        qty = num(r[6] if len(r) > 6 else None)
        if not sku:
            continue
        if qty is None:
            continue
        if qty <= 0:
            continue
        lines.append(
            {
                "workbook_line_key": f"{sheet_name}!r{i + 1}",
                "sku": sku.upper(),
                "description": txt(r[3] if len(r) > 3 else None),
                "qty_total": qty,
                "source_row": i + 1,
            }
        )

    # unit mix below subtotal — label in col C, count in col D (product col)
    for i in range(header_idx + 1, len(rows)):
        r = list(rows[i] or [])
        label = txt(r[2] if len(r) > 2 else None)
        count = num(r[3] if len(r) > 3 else None)
        if label and count is not None and len(label) <= 12 and re.match(r"^[A-Za-z0-9]", label):
            unit_mix.append({"label": label, "count": int(count)})

    return {"header": header, "lines": lines, "unit_mix": unit_mix}


def parse_asia_po_form(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = -1
    for i, row in enumerate(rows[:40]):
        vals = [txt(c) for c in (row or [])]
        if "SKU NUMBER" in vals and "QTY ORDER" in vals:
            header_idx = i
            break
    if header_idx < 0:
        raise ValueError("asia_po: SKU NUMBER / QTY ORDER header not found on Form sheet")

    header = {"project_name": None, "quote_sheet": "Form"}
    for row in rows[:header_idx]:
        r = list(row or [])
        ref = txt(r[1] if len(r) > 1 else None) or txt(r[2] if len(r) > 2 else None)
        if ref and ref.upper().startswith("REFERENCE:"):
            header["project_name"] = ref.split(":", 1)[-1].strip()

    lines = []
    for i in range(header_idx + 1, len(rows)):
        r = list(rows[i] or [])
        sku = txt(r[1] if len(r) > 1 else None)
        qty = num(r[3] if len(r) > 3 else None)
        if not sku or not qty or qty <= 0:
            if sku and txt(r[3] if len(r) > 3 else None) == "TOTAL COST:":
                break
            continue
        lines.append(
            {
                "workbook_line_key": f"Form!r{i + 1}",
                "sku": sku.upper(),
                "description": txt(r[2] if len(r) > 2 else None),
                "qty_total": qty,
                "source_row": i + 1,
            }
        )
    return {"header": header, "lines": lines, "unit_mix": []}


SKIP_INVOICE_SKUS = {
    "MISC",
    "MISC-PARTS",
    "SF-INS",
    "SF-FR",
    "TEXT",
    "SPARE-PARTS",
    "AVATAX",
    "FIXED",
    "UOM",
    "FREIGHT",
    "INSTALLATION (EXCLUDES EXTRACTION,",
}


def parse_sage_invoice_pdf(path: Path) -> dict:
    proc = subprocess.run(
        ["pdftotext", str(path), "-"],
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        raise ValueError(f"pdftotext failed: {proc.stderr[:200]}")
    raw_lines = [l.strip() for l in proc.stdout.splitlines() if l.strip()]

    inv = None
    for s in raw_lines:
        m = re.match(r"^(\d{2}-\d{3}-I)$", s)
        if m:
            inv = m.group(1)
            break

    inline = re.compile(r"^([\d,]+)\s+([A-Z0-9][A-Z0-9._-]{2,})$")
    sku_only = re.compile(r"^([A-Z0-9][A-Z0-9._-]{2,})$")
    qty_only = re.compile(r"^([\d,]+)$")

    def is_sku_token(s: str) -> bool:
        if s in SKIP_INVOICE_SKUS or s.startswith("SF-"):
            return False
        if re.match(r"^[\d,]+$", s):
            return False
        if not sku_only.match(s):
            return False
        if s.endswith(" EA"):
            return False
        return True

    merged: dict[str, dict] = {}

    def push(sku: str, qty: float, key: str) -> None:
        sku = sku.upper().strip()
        if not is_sku_token(sku):
            return
        if qty <= 0:
            return
        prev = merged.get(sku)
        row = {
            "workbook_line_key": key,
            "sku": sku,
            "description": None,
            "qty_total": qty,
            "source_row": None,
        }
        if not prev or qty > prev["qty_total"]:
            merged[sku] = row

    # Pass 1: inline "1,472 VN1618-..." lines anywhere in doc
    for i, s in enumerate(raw_lines, 1):
        m = inline.match(s)
        if m:
            push(m.group(2), float(m.group(1).replace(",", "")), f"pdf:L{i}")

    # Pass 2: qty ord triplets followed by SKU list (fixed-style layout)
    i = 0
    while i < len(raw_lines):
        if raw_lines[i] == "Ord." and i > 0 and raw_lines[i - 1] == "Qty.":
            i += 1
            while i < len(raw_lines) and raw_lines[i] != "B/O":
                i += 1
            i += 1  # past B/O column header
            nums: list[float] = []
            while i < len(raw_lines):
                s = raw_lines[i]
                if s in ("Item Number", "Customer No.", "Description") or is_sku_token(s):
                    break
                m = inline.match(s)
                if m:
                    push(m.group(2), float(m.group(1).replace(",", "")), f"pdf:block:L{i+1}")
                    i += 1
                    continue
                if qty_only.match(s):
                    nums.append(float(s.replace(",", "")))
                i += 1
            ord_qtys = nums[0::3] if nums else []
            skus: list[str] = []
            while i < len(raw_lines):
                s = raw_lines[i]
                if s in ("Item Number", "Order Date", "Customer No.", "Description", "Ship Via"):
                    break
                m = inline.match(s)
                if m:
                    push(m.group(2), float(m.group(1).replace(",", "")), f"pdf:block:L{i+1}")
                    i += 1
                    continue
                if is_sku_token(s):
                    skus.append(s.upper())
                i += 1
            for q, sku in zip(ord_qtys, skus):
                push(sku, q, f"pdf:zip:{sku}")
            continue
        i += 1

    # Pass 3: description block qty on line above SKU (loose invoice page 2)
    pending: float | None = None
    for i, s in enumerate(raw_lines, 1):
        if qty_only.match(s) and not inline.match(s):
            pending = float(s.replace(",", ""))
            continue
        if pending is not None and is_sku_token(s):
            push(s, pending, f"pdf:pair:L{i}")
            pending = None
            continue
        if not qty_only.match(s):
            pending = None

    return {
        "header": {"project_name": inv, "quote_sheet": "sage_invoice_pdf"},
        "lines": list(merged.values()),
        "unit_mix": [],
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", help="Force sheet name")
    ap.add_argument("--format", choices=["auto", "quote_rev", "asia_po", "sage_invoice_pdf"], default="auto")
    args = ap.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        print(json.dumps({"error": f"file not found: {path}"}))
        sys.exit(1)

    warnings: list[str] = []
    fmt = args.format
    parsed = None

    if fmt == "sage_invoice_pdf" or (fmt == "auto" and path.suffix.lower() == ".pdf"):
        parsed = parse_sage_invoice_pdf(path)
        fmt = "sage_invoice_pdf"
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if args.sheet:
                ws = wb[args.sheet]
                if fmt == "asia_po" or args.sheet.lower() == "form":
                    parsed = parse_asia_po_form(ws)
                    fmt = "asia_po"
                else:
                    parsed = parse_quote_sheet(ws, args.sheet)
                    fmt = "quote_rev"
            elif fmt == "asia_po" or (fmt == "auto" and "Form" in wb.sheetnames):
                parsed = parse_asia_po_form(wb["Form"])
                fmt = "asia_po"
            else:
                sheets = find_quote_sheets(wb.sheetnames)
                if not sheets:
                    raise ValueError(
                        f"No Quote Rev / Quote - sheet in {wb.sheetnames}. "
                        "Leo SO TABLE xlsm is enterprise inventory — not a CSL Sales workbook."
                    )
                all_lines = []
                unit_mix = []
                header = None
                for qsn in sheets:
                    part = parse_quote_sheet(wb[qsn], qsn)
                    header = header or part["header"]
                    all_lines.extend(part["lines"])
                    unit_mix.extend(part.get("unit_mix") or [])
                parsed = {"header": header or {}, "lines": all_lines, "unit_mix": unit_mix}
                fmt = "quote_rev"
        finally:
            wb.close()

    out = {
        "source": str(path),
        "format": fmt,
        "header": parsed["header"],
        "lines": parsed["lines"],
        "unit_mix": parsed.get("unit_mix") or [],
        "line_count": len(parsed["lines"]),
        "warnings": warnings,
    }
    print(json.dumps(out))


if __name__ == "__main__":
    main()
