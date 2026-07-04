#!/usr/bin/env python3
"""Extract Troubadour (25019/25198) unit matrix from Box xlsx → JSON for Registry-iQ ingest."""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

BOX_MATRIX = Path(
    "/Users/geoffreyjackson/Library/CloudStorage/Box-Box/Team Folder/Projects/"
    "25019-Lubbock, TX - 14th Street SH/PROJECT MANAGING/SHOP DRAWINGS/"
    "Unit & Shop Drawing Matrix Troubadour.xlsx"
)
OUT = Path(__file__).resolve().parent.parent / ".firecrawl" / "troubadour-matrix.json"


def norm_unit(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() == "unit # ":
        return None
    try:
        if isinstance(val, (int, float)) and float(val).is_integer():
            return str(int(float(val)))
    except (TypeError, ValueError):
        pass
    return s


def infer_beds(unit_type: str | None) -> int:
    if not unit_type:
        return 0
    t = unit_type.strip().upper()
    m = re.match(r"^B(\d+)", t)
    if m:
        return int(m.group(1))
    m = re.match(r"^D(\d+(?:\.\d+)?)", t)
    if m:
        return int(float(m.group(1)))
    if t.startswith("S"):
        return 0
    if t.startswith("A"):
        return 1
    if t.startswith("E"):
        return 1
    return 0


def main() -> None:
    if not BOX_MATRIX.exists():
        raise SystemExit(f"Matrix not found: {BOX_MATRIX}")

    wb = openpyxl.load_workbook(BOX_MATRIX, read_only=True, data_only=True)
    ws = wb["Lubbock"]
    headers = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]
    col_idx = {str(h).strip(): i for i, h in enumerate(headers) if h}

    phase_by_unit: dict[str, str | None] = {}
    if "SHOWERS Sequence" in wb.sheetnames:
        sws = wb["SHOWERS Sequence"]
        for row in sws.iter_rows(min_row=4, values_only=True):
            u = norm_unit(row[0])
            if u:
                phase_by_unit[u] = str(row[4]).strip() if row[4] is not None else None

    units = []
    type_counts: Counter[str] = Counter()
    drawing_counts: Counter[str] = Counter()

    for row in ws.iter_rows(min_row=4, values_only=True):
        unit = norm_unit(row[0])
        if not unit:
            continue
        unit_type = str(row[2]).strip() if row[2] else None
        area = str(row[3]).strip() if row[3] is not None else None
        thus_opp = str(row[4]).strip() if row[4] is not None else None
        level = row[1]
        kitchen_cab = str(row[5]).strip() if row[5] else None

        drawings = {}
        for label in (
            "Kitchen Cab",
            "KITCHEN TOP SD",
            "BATH TOP SD",
            "BATH 2 TOP SD",
            "BATH 3 TOP SD2",
            "BATH 4 TOP SD3",
            "BATH 5 TOP SD4",
        ):
            idx = col_idx.get(label)
            if idx is None:
                continue
            val = row[idx] if idx < len(row) else None
            if val:
                drawings[label] = str(val).strip()

        if kitchen_cab:
            drawing_counts[kitchen_cab] += 1
        if unit_type:
            type_counts[unit_type] += 1

        units.append(
            {
                "unit_number": unit,
                "level": int(level) if isinstance(level, (int, float)) and float(level).is_integer() else level,
                "unit_type": unit_type,
                "construction_area": area,
                "thus_opp": thus_opp,
                "phase_no": phase_by_unit.get(unit),
                "kitchen_cab": kitchen_cab,
                "drawings": drawings,
            }
        )

    unit_types = []
    for name, count in sorted(type_counts.items(), key=lambda x: (-x[1], x[0])):
        unit_types.append(
            {
                "unit_type_name": name,
                "unit_count": count,
                "beds_per_unit": infer_beds(name),
                "standard_bedrooms": infer_beds(name),
            }
        )

    payload = {
        "source": str(BOX_MATRIX),
        "property_key": "troubadour_14th_street_lubbock",
        "unit_count": len(units),
        "unit_type_count": len(unit_types),
        "kitchen_cab_variants": len(drawing_counts),
        "units": units,
        "unit_types": unit_types,
        "kitchen_cab_counts": dict(drawing_counts.most_common()),
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} — {len(units)} units, {len(unit_types)} types, {len(drawing_counts)} kitchen cab codes")


if __name__ == "__main__":
    main()
